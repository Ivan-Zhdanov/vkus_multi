import openai
from openpyxl import load_workbook
import time


def list_api():
    list_api = []
    wb = load_workbook('API_KEYS.xlsx')
    ws = wb.active
    for row in range(1, ws.max_row):
        if ws[f'B{row}'].value < 1000:
            print(f'Взятый API {row}', ws[f'A{row}'].value)
            count = ws[f'B{row}'].value
            ws[f'B{row}'] = count + 1
            API_KEY = ws[f'A{row}'].value
            list_api.append(API_KEY)
            wb.save('API_KEYS.xlsx')
    return list_api

def api_cheker():

    # API_KEY = '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'
    print(time.time())
    wb = load_workbook('API_KEYS.xlsx')
    ws = wb.active

    # вызываем запрос

    # API_KEY = None
    # while API_KEY == None:
    while True:
        for row in range(1, ws.max_row):
            # print('НОМЕР СТРОКИ', row)
            if ws[f'B{row}'].value < 1000:
                if int(time.time()) - ws[f'C{row}'].value > 45:
                    print(f'Взятый API {row}', ws[f'A{row}'].value)

                    # запись
                    count = ws[f'B{row}'].value
                    ws[f'B{row}'] = count + 1
                    ws[f'C{row}'] = int(time.time())
                    API_KEY =  ws[f'A{row}'].value
                    wb.save('API_KEYS.xlsx')
                    return API_KEY
            # print('следующий API')
        time.sleep(20)
